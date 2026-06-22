"""
Джарвис v13 — капитальная перестройка.
Архитектура: БЕЗ ConversationHandler. Все кнопки — глобальные callbacks.
Состояния — через context.user_data["mode"]. Кнопки работают из любого места.

Правила:
- ТЗ -> сразу Word; Критерии -> сразу Word; Анализ -> сразу Word
- Письма -> сразу Word (всегда)
- Переговоры -> Word или PDF на выбор
- После каждого документа: "Всё отлично" / "Есть замечания" -> правки той же версии
- Файл в чате: Проанализировать / Отредактировать / Написать ответ
- ТЗ: только 3 направления. Критерии: 3 направления + Свой вариант
"""

import os, json, logging, base64, tempfile
import anthropic
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes,
)
from agent import TenderAgent, QUESTIONS
from voice_handler import transcribe_voice

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

sessions: dict[int, TenderAgent] = {}
last_doc: dict[int, dict] = {}
bot_msgs: dict[int, list] = {}

claude = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
WEB_SEARCH_TOOL = {"type": "web_search_20250305", "name": "web_search"}

# ─── Доступ ──────────────────────────────────────────────────────────────────
USERS_FILE = "/tmp/bot_users.json"
DYNAMIC_FILE = "/tmp/dynamic_users.json"
ALLOWED_USERS: set = set()
for _s in os.environ.get("ALLOWED_USERS", "").split(","):
    try: ALLOWED_USERS.add(int(_s.strip()))
    except ValueError: pass

def _load_dynamic():
    try:
        if os.path.exists(DYNAMIC_FILE):
            return set(json.load(open(DYNAMIC_FILE)).get("users", []))
    except: pass
    return set()

def _save_dynamic(s):
    try: json.dump({"users": list(s)}, open(DYNAMIC_FILE, "w"))
    except Exception as e: logger.error(e)

DYNAMIC_USERS: set = _load_dynamic()

def is_allowed(uid):
    if not ALLOWED_USERS and not DYNAMIC_USERS: return True
    return uid in ALLOWED_USERS or uid in DYNAMIC_USERS

def save_user(uid, username=""):
    try:
        data = json.load(open(USERS_FILE)) if os.path.exists(USERS_FILE) else {}
        data[str(uid)] = username
        json.dump(data, open(USERS_FILE, "w"))
    except Exception as e: logger.error(e)

def get_all_users():
    try:
        if os.path.exists(USERS_FILE): return json.load(open(USERS_FILE))
    except: pass
    return {}

def remember(uid, text):
    bot_msgs.setdefault(uid, []).append(text)
    if len(bot_msgs[uid]) > 20: bot_msgs[uid] = bot_msgs[uid][-20:]

# ─── Промпты ─────────────────────────────────────────────────────────────────
CHAT_SYSTEM = """Ты - Джарвис, умный и немного саркастичный помощник. Как Джарвис из Железного человека: чёткий, профессиональный, с лёгкой иронией. Помогаешь всегда, без лишней воды.

Твоя ОСНОВНАЯ ЭКСПЕРТИЗА - коммерческие закупки работ и услуг (не государственные закупки, не закупки товаров). В этой теме ты разбираешься глубоко и профессионально.

Но ты свободно общаешься и на любые другие темы - шутки, новости, общие вопросы, что угодно. Специализация - это твоя сильная сторона, а не ограничение на то, о чём можно говорить. Если спрашивают что-то не по закупкам - отвечай нормально, без отказов и без необходимости "возвращать к теме".

Используй поиск в интернете когда это полезно - для новостей, актуальных фактов, любых вопросов требующих свежей информации.

Помни весь контекст разговора включая фото.
Смайлики активно. Русский язык."""

REVIEW_SYSTEM = ("Эксперт по тендерам и деловым документам. Внеси правки и верни ПОЛНЫЙ исправленный текст документа. "
                  "Только текст, без комментариев.\n\n"
                  "Если в задании просят НАЙТИ, ОТМЕТИТЬ или ВЫДЕЛИТЬ конкретные фрагменты текста "
                  "(повторы, ошибки, рискованные формулировки и т.п.) - оборачивай именно эти фрагменты "
                  "маркером ==текст== - они будут подсвечены жёлтым в документе. "
                  "Если задание не про выделение - не используй этот маркер вообще.")

NEGOTIATION_SYSTEM = """Эксперт по закупочным переговорам. Цель: снижение цены и улучшение условий.
Сценарий без воды:
1. ПОЗИЦИЯ ЗАКУПЩИКА
2. ОТКРЫТИЕ (2-3 варианта первых фраз)
3. АРГУМЕНТЫ для давления на цену
4. ВОЗРАЖЕНИЯ участника и точные ответы
5. ЗАКРЫТИЕ сделки
Только конкретные фразы. Русский язык."""

LETTER_SYSTEM = """Составляешь официальные деловые письма на русском языке.
Профессионально, структурированно, вежливо.
Начни с обращения, изложи суть, закончи подписью.
Только текст письма."""

ANALYSIS_SYSTEM = """Ты опытный эксперт по коммерческим закупкам работ и услуг.
Проводишь экспертизу документов организатора коммерческого тендера.
Анализируй по критериям:
1. Ошибки и противоречия
2. Ограничивающие требования (риск претензий)
3. Нечёткие формулировки
4. Недостающие требования
5. Конкретные предложения по исправлению
Указывай разделы документа, предлагай готовые формулировки. Русский язык."""

VOICE_EXTRACT_SYSTEM = """Ты помощник по коммерческим закупкам.
Пользователь надиктовал голосовое сообщение. Извлеки намерение и параметры.

Верни JSON строго в формате:
{"intent": "tz" | "criteria" | "negotiation" | "letter" | "analysis" | "chat",
 "direction": "cleaning" | "it" | "repair" | null,
 "params": {"вопрос": "ответ"},
 "original_task": "краткое описание задачи"}

Если параметр не упомянут - не включай его. Только JSON."""

NEGOTIATION_STEPS = [
    {"q": "Что закупаем?", "opts": ["Клининговые услуги", "IT-услуги", "Ремонт оборудования", "Строительные работы", "Консалтинг / аудит", "Другие услуги"], "free": True},
    {"q": "Кто придёт от участника?", "opts": ["Директор/собственник", "Коммерческий директор", "Менеджер по продажам", "Неизвестно"], "free": False},
    {"q": "НМЦ (начальная цена)?", "opts": ["До 1 млн руб.", "1-5 млн руб.", "5-20 млн руб.", "Более 20 млн руб."], "free": True},
    {"q": "На сколько снижаем цену?", "opts": ["На 5-10%", "На 10-20%", "На 20-30%", "Максимально"], "free": False},
    {"q": "Есть альтернативные участники?", "opts": ["Да, 2+ конкурента", "Есть 1 альтернатива", "Нет, единственный"], "free": False},
    {"q": "Доп. цели переговоров?", "opts": ["Только снижение цены", "Цена + сроки", "Цена + гарантии", "Цена + объём работ"], "free": False},
]

# ─── Утилиты ─────────────────────────────────────────────────────────────────
def set_mode(context, mode):
    context.user_data["mode"] = mode

def get_mode(context):
    return context.user_data.get("mode")

async def get_text(update):
    if update.message.text: return update.message.text.strip()
    if update.message.voice:
        await update.message.reply_text("Слушаю...")
        file = await update.message.voice.get_file()
        data = await file.download_as_bytearray()
        text = await transcribe_voice(bytes(data))
        await update.message.reply_text("Услышал: " + text[:150])
        return text
    return None

async def get_image_b64(update):
    photo = None
    if update.message.photo:
        photo = update.message.photo[-1]
    elif update.message.document and update.message.document.mime_type and \
         update.message.document.mime_type.startswith("image/"):
        photo = update.message.document
    if not photo: return None
    file = await photo.get_file()
    data = await file.download_as_bytearray()
    b64 = base64.standard_b64encode(bytes(data)).decode()
    mime = getattr(photo, "mime_type", None) or "image/jpeg"
    return b64, mime

def _read_docx(path):
    """Читает docx ВКЛЮЧАЯ таблицы."""
    from docx import Document as D
    d = D(path)
    parts = []
    for p in d.paragraphs:
        if p.text.strip(): parts.append(p.text)
    for table in d.tables:
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells]
            if any(cells):
                parts.append(" | ".join(cells))
    return "\n".join(parts)

async def extract_doc_text(update):
    """Читает Word/PDF/txt, включая таблицы в docx."""
    doc = update.message.document
    if not doc: return None
    fname = doc.file_name or ""
    mime = doc.mime_type or ""
    logger.info("Reading doc: " + fname + " mime=" + mime)
    file = await doc.get_file()
    data = bytes(await file.download_as_bytearray())
    fl = fname.lower()
    suffix = os.path.splitext(fname)[1] or ".bin"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(data); tmp.close()
    try:
        if fl.endswith(".docx") or "wordprocessingml" in mime or ("word" in mime and not fl.endswith(".doc")):
            return _read_docx(tmp.name)
        if fl.endswith(".pdf") or "pdf" in mime:
            from pypdf import PdfReader
            return "\n".join(p.extract_text() or "" for p in PdfReader(tmp.name).pages).strip()
        if fl.endswith(".txt") or "text" in mime:
            return data.decode("utf-8", errors="replace").strip()
        if fl.endswith(".xlsx") or fl.endswith(".xls") or "spreadsheet" in mime or "excel" in mime:
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                parts.append("=== Лист: " + ws.title + " ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        parts.append(" | ".join(cells))
            wb.close()
            return "\n".join(parts)
        if fl.endswith(".doc"):
            return None  # старый формат не поддерживаем
        # octet-stream или неизвестный mime — пробуем по очереди
        try:
            return _read_docx(tmp.name)
        except Exception:
            pass
        try:
            import openpyxl as _opx
            wb2 = _opx.load_workbook(tmp.name, read_only=True, data_only=True)
            parts2 = []
            for ws2 in wb2.worksheets:
                parts2.append("=== Лист: " + ws2.title + " ===")
                for row2 in ws2.iter_rows(values_only=True):
                    cells2 = [str(c) if c is not None else "" for c in row2]
                    if any(c.strip() for c in cells2):
                        parts2.append(" | ".join(cells2))
            wb2.close()
            if parts2: return "\n".join(parts2)
        except Exception:
            pass
        try:
            from pypdf import PdfReader
            t = "\n".join(p.extract_text() or "" for p in PdfReader(tmp.name).pages).strip()
            if t: return t
        except Exception:
            pass
        try:
            return data.decode("utf-8", errors="replace").strip()
        except Exception:
            pass
    except Exception as e:
        logger.error("Doc read: " + str(e), exc_info=True)
    finally:
        try: os.remove(tmp.name)
        except: pass
    return None

async def send_file(context, chat_id, path, filename, caption):
    with open(path, "rb") as fh:
        await context.bot.send_document(chat_id=chat_id, document=fh, filename=filename, caption=caption)
    try: os.remove(path)
    except: pass

async def send_review(context, chat_id):
    await context.bot.send_message(chat_id=chat_id, text="Всё устраивает?", reply_markup=review_kb())

async def send_q(msg, result):
    text = result["question"]; opts = result.get("options", [])
    if opts:
        kb = [[InlineKeyboardButton(o, callback_data="ans_" + str(i))] for i, o in enumerate(opts)]
        kb.append([InlineKeyboardButton("Свой вариант", callback_data="ans_custom")])
        await msg.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await msg.reply_text(text)

# ─── Клавиатуры ──────────────────────────────────────────────────────────────
def nav_kb():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("/new — Новый запрос"), KeyboardButton("/cancel — Отмена"), KeyboardButton("/help — Помощь")]],
        resize_keyboard=True, is_persistent=True)

def menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 Техническое задание", callback_data="menu_tz")],
        [InlineKeyboardButton("📋 Критерии допуска", callback_data="menu_criteria")],
        [InlineKeyboardButton("🤝 Сценарий переговоров", callback_data="menu_negotiation")],
        [InlineKeyboardButton("✉️ Написать письмо", callback_data="menu_letter")],
        [InlineKeyboardButton("🔎 Анализ документа", callback_data="menu_analysis")],
        [InlineKeyboardButton("⚖️ Сравнить два документа", callback_data="menu_compare")],
    ])

def dir_kb(doc_type="tz_only"):
    rows = [
        [InlineKeyboardButton("🧹 Клининговые услуги", callback_data="dir_cleaning")],
        [InlineKeyboardButton("💻 IT-услуги и автоматизация", callback_data="dir_it")],
        [InlineKeyboardButton("🔧 Ремонт и техобслуживание", callback_data="dir_repair")],
        [InlineKeyboardButton("✏️ Свой вариант", callback_data="dir_custom")],
    ]
    return InlineKeyboardMarkup(rows)

def hasdoc_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Да, загружу документ", callback_data="hasdoc_yes")],
        [InlineKeyboardButton("Нет, задавай вопросы", callback_data="hasdoc_no")],
    ])

def review_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Всё отлично!", callback_data="review_ok")],
        [InlineKeyboardButton("✏️ Есть замечания", callback_data="review_edit")],
    ])

def word_pdf_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 Word", callback_data="fmt_docx")],
        [InlineKeyboardButton("📕 PDF", callback_data="fmt_pdf")],
        [InlineKeyboardButton("📄+📕 Оба", callback_data="fmt_both")],
    ])

def neg_kb(step):
    s = NEGOTIATION_STEPS[step]
    btns = [[InlineKeyboardButton(o, callback_data="neg_" + str(step) + "_" + str(i))] for i, o in enumerate(s["opts"])]
    if s["free"]: btns.append([InlineKeyboardButton("Свой вариант", callback_data="neg_" + str(step) + "_custom")])
    return InlineKeyboardMarkup(btns)

def letter_type_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📩 Ответ на письмо", callback_data="letter_reply")],
        [InlineKeyboardButton("📝 Новое письмо", callback_data="letter_new")],
    ])

def doc_action_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔎 Проанализировать", callback_data="docact_analyze")],
        [InlineKeyboardButton("✏️ Отредактировать", callback_data="docact_edit")],
        [InlineKeyboardButton("✉️ Написать ответ", callback_data="docact_reply")],
    ])

def photo_action_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✉️ Написать ответ", callback_data="photo_act_reply")],
        [InlineKeyboardButton("✏️ Отредактировать", callback_data="photo_act_edit")],
        [InlineKeyboardButton("🔎 Проанализировать", callback_data="photo_act_analyze")],
    ])

# ─── Распознавание намерений ─────────────────────────────────────────────────
def detect_intent(text):
    tl = text.lower().strip()
    if any(w in tl for w in ["проверь", "проанализируй", "экспертиза", "найди ошибки", "разбери", "анализ документа", "анализ тз", "сделай анализ"]):
        return "menu_analysis"
    if any(w in tl for w in ["сравни", "сравнить", "сравнение", "отличия между", "разница между"]):
        return "menu_compare"
    tz_words = ["тз", "техзадание", "техническое задание"]
    actions = ["нужно", "нужен", "нужны", "хочу", "сделай", "составь", "напиши", "подготовь", "создай"]
    if any(w in tl for w in tz_words):
        if any(a in tl for a in actions) or len(tl.split()) <= 3:
            return "menu_tz"
    if "критери" in tl:
        if any(a in tl for a in actions) or len(tl.split()) <= 3:
            return "menu_criteria"
    if "переговор" in tl or "скрипт" in tl:
        return "menu_negotiation"
    if "письм" in tl:
        if any(a in tl for a in actions) or "ответ на" in tl:
            return "menu_letter"
    return None

# ─── Команды ─────────────────────────────────────────────────────────────────
async def cmd_start(update, context):
    uid = update.effective_user.id
    if not is_allowed(uid):
        await update.message.reply_text("Нет доступа. Обратись к администратору."); return
    save_user(uid, update.effective_user.username or "")
    sessions.pop(uid, None); context.user_data.clear()
    await update.message.reply_text(
        "Привет! Я Джарвис — ваш персональный помощник по коммерческим закупкам.\n\n"
        "Нажми /new чтобы создать документ или просто напиши что нужно!",
        reply_markup=nav_kb())

async def cmd_new(update, context):
    uid = update.effective_user.id
    if not is_allowed(uid):
        await update.message.reply_text("Нет доступа."); return
    save_user(uid, update.effective_user.username or "")
    sessions.pop(uid, None); last_doc.pop(uid, None); context.user_data.clear()
    await update.message.reply_text("Что делаем?", reply_markup=menu_kb())

async def cmd_cancel(update, context):
    uid = update.effective_user.id
    sessions.pop(uid, None); last_doc.pop(uid, None); context.user_data.clear()
    await update.message.reply_text("Отменено. /new для нового запроса.", reply_markup=nav_kb())

async def cmd_help(update, context):
    await update.message.reply_text(
        "Возможности:\n\n"
        "/new - создать документ (ТЗ, критерии, переговоры, письмо, анализ)\n"
        "Напиши 'нужно ТЗ', 'критерии', 'переговоры' - сразу запущу режим\n"
        "Загрузи файл - спрошу что с ним сделать\n"
        "Отправь фото письма - предложу действия\n"
        "Голосовое сообщение - пойму и запущу нужный режим\n"
        "'сохрани' - сохраню последний ответ в Word\n"
        "/cancel - отменить текущую операцию",
        reply_markup=nav_kb())

async def cmd_users(update, context):
    uid = update.effective_user.id
    if uid != int(os.environ.get("ADMIN_ID", "0")):
        await update.message.reply_text("Только для администратора."); return
    all_users = get_all_users(); whitelist = ALLOWED_USERS | DYNAMIC_USERS
    lines = ["Пользователи:\n"]
    for uid_str, uname in all_users.items():
        s = "OK" if int(uid_str) in whitelist else "NO"
        lines.append("[" + s + "] " + uid_str + " - @" + (uname or "-"))
    lines.append("\nВ белом списке: " + str(len(whitelist)))
    await update.message.reply_text("\n".join(lines))

async def cmd_adduser(update, context):
    if update.effective_user.id != int(os.environ.get("ADMIN_ID", "0")):
        await update.message.reply_text("Только для администратора."); return
    if not context.args:
        await update.message.reply_text("Укажи ID: /adduser 123456789"); return
    try: new_uid = int(context.args[0])
    except: await update.message.reply_text("Неверный ID."); return
    DYNAMIC_USERS.add(new_uid); _save_dynamic(DYNAMIC_USERS)
    try:
        await context.bot.send_message(chat_id=new_uid, text="Тебе открыт доступ! Напиши /start")
        notified = "Уведомлён."
    except: notified = "Не смог уведомить."
    await update.message.reply_text("Пользователь " + str(new_uid) + " добавлен. " + notified)

async def cmd_removeuser(update, context):
    if update.effective_user.id != int(os.environ.get("ADMIN_ID", "0")):
        await update.message.reply_text("Только для администратора."); return
    if not context.args:
        await update.message.reply_text("Укажи ID: /removeuser 123456789"); return
    try: rem_uid = int(context.args[0])
    except: await update.message.reply_text("Неверный ID."); return
    if rem_uid in DYNAMIC_USERS:
        DYNAMIC_USERS.discard(rem_uid); _save_dynamic(DYNAMIC_USERS)
        await update.message.reply_text("Пользователь " + str(rem_uid) + " удалён.")
    elif rem_uid in ALLOWED_USERS:
        await update.message.reply_text(str(rem_uid) + " в ALLOWED_USERS на Railway - удали там вручную.")
    else:
        await update.message.reply_text(str(rem_uid) + " не найден.")

async def cmd_broadcast(update, context):
    if update.effective_user.id != int(os.environ.get("ADMIN_ID", "0")):
        await update.message.reply_text("Только для администратора."); return
    text = " ".join(context.args) if context.args else ""
    if not text:
        await update.message.reply_text("Напиши:\n/broadcast текст"); return
    users = get_all_users(); sent = failed = 0
    await update.message.reply_text("Рассылаю " + str(len(users)) + " пользователям...")
    for uid_str in users:
        try:
            await context.bot.send_message(chat_id=int(uid_str), text=text); sent += 1
        except: failed += 1
    await update.message.reply_text("Отправлено: " + str(sent) + ", не доставлено: " + str(failed))

# ─── Меню (callbacks) ────────────────────────────────────────────────────────
async def cb_menu(update, context):
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    if not is_allowed(uid):
        await q.edit_message_text("Нет доступа."); return
    sessions.pop(uid, None)
    context.user_data.clear()

    if q.data == "menu_negotiation":
        context.user_data["neg_answers"] = {}; context.user_data["neg_step"] = 0
        set_mode(context, "negotiation")
        await q.edit_message_text(NEGOTIATION_STEPS[0]["q"], reply_markup=neg_kb(0))
        return

    if q.data == "menu_letter":
        await q.edit_message_text("Что нужно?", reply_markup=letter_type_kb())
        return

    if q.data == "menu_analysis":
        set_mode(context, "awaiting_analysis_doc")
        await q.edit_message_text(
            "Загрузи документ для экспертизы (Word, PDF, txt, фото или вставь текст).\n"
            "Найду ошибки, противоречия, завышенные требования и дам предложения.")
        return

    if q.data == "menu_compare":
        set_mode(context, "awaiting_compare_doc1")
        await q.edit_message_text(
            "Загрузи первый документ (Word, PDF, Excel, txt):")
        return

    doc_map = {"menu_tz": "tz_only", "menu_criteria": "criteria_only"}
    doc_type = doc_map.get(q.data, "tz_only")
    context.user_data["doc_type"] = doc_type
    await q.edit_message_text("Выбери направление:", reply_markup=dir_kb(doc_type))

async def cb_direction(update, context):
    q = update.callback_query; await q.answer()
    direction = q.data.replace("dir_", "")
    doc_type = context.user_data.get("doc_type", "tz_only")
    if direction == "custom":
        if doc_type == "criteria_only":
            # Для критериев — старый флоу: просто название направления
            set_mode(context, "awaiting_custom_direction")
            await q.edit_message_text(
                "Напиши направление закупки своими словами\n"
                "(например: охрана объектов, вывоз мусора, обслуживание лифтов):")
        else:
            # Для ТЗ — новый флоу: свободная генерация с загрузкой примеров
            set_mode(context, "awaiting_free_tz_name")
            await q.edit_message_text(
                "Напиши что именно закупаем — название или краткое описание:\n"
                "(например: создание имиджевых буклетов, охрана объектов, вывоз мусора)")
        return
    context.user_data["direction"] = direction
    await q.edit_message_text(
        "Есть документ от заказчика? (ТЗ-черновик, письмо)\n"
        "Если да - загружу и задам только недостающие вопросы.",
        reply_markup=hasdoc_kb())

async def cb_hasdoc(update, context):
    q = update.callback_query; await q.answer()
    if q.data == "hasdoc_yes":
        set_mode(context, "awaiting_customer_doc")
        await q.edit_message_text("Загрузи документ (Word, PDF или txt):")
        return
    await q.edit_message_text("Хорошо, задам вопросы.")
    await _start_questions(q.message.chat_id, update, context)

async def _start_questions(chat_id, update, context):
    direction = context.user_data.get("direction", "cleaning")
    doc_type = context.user_data.get("doc_type", "tz_only")
    custom_name = context.user_data.get("custom_direction_name", "")
    agent = TenderAgent(direction=direction, doc_type=doc_type, custom_name=custom_name)
    sessions[update.effective_user.id] = agent
    set_mode(context, "answering")
    result = await agent.get_next_question()
    text = result["question"]; opts = result.get("options", [])
    if opts:
        kb = [[InlineKeyboardButton(o, callback_data="ans_" + str(i))] for i, o in enumerate(opts)]
        kb.append([InlineKeyboardButton("Свой вариант", callback_data="ans_custom")])
        await context.bot.send_message(chat_id=chat_id, text=text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await context.bot.send_message(chat_id=chat_id, text=text)

# ─── Вопросы агента ──────────────────────────────────────────────────────────
async def cb_answer(update, context):
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    agent = sessions.get(uid)
    if not agent:
        await q.edit_message_text("Сессия устарела. /new"); return
    if q.data == "ans_custom":
        await q.edit_message_text(q.message.text + "\n\nВведи свой вариант:")
        return
    idx = int(q.data.replace("ans_", ""))
    opts = agent.last_question.get("options", [])
    answer = opts[idx] if idx < len(opts) else ""
    await q.edit_message_text(agent.last_question["question"] + "\n> " + answer)
    await _handle_answer(uid, q.message.chat_id, update, context, answer)

async def _handle_answer(uid, chat_id, update, context, answer):
    agent = sessions[uid]
    prefilled = context.user_data.get("prefilled", {})
    if prefilled and not any(a.get("prefilled") for a in agent.answers):
        for fq, fa in prefilled.items():
            agent.answers.append({"question": fq, "answer": fa, "prefilled": True})
    result = await agent.submit_answer(answer)
    if result["status"] == "question":
        text = result["question"]; opts = result.get("options", [])
        if opts:
            kb = [[InlineKeyboardButton(o, callback_data="ans_" + str(i))] for i, o in enumerate(opts)]
            kb.append([InlineKeyboardButton("Свой вариант", callback_data="ans_custom")])
            await context.bot.send_message(chat_id=chat_id, text=text, reply_markup=InlineKeyboardMarkup(kb))
        else:
            await context.bot.send_message(chat_id=chat_id, text=text)
        return
    await context.bot.send_message(chat_id=chat_id, text="Данные собраны! Генерирую...")
    await do_generate(uid, chat_id, context)

# ─── Генерация ТЗ и критериев ────────────────────────────────────────────────
async def do_generate(uid, chat_id, context):
    from docx_generator import generate_tz_docx
    agent = sessions.get(uid)
    if not agent:
        await context.bot.send_message(chat_id=chat_id, text="Сессия устарела. /new"); return
    try:
        if agent.doc_type in ("tz_only", "both"):
            await context.bot.send_message(chat_id=chat_id, text="Генерирую ТЗ...")
            content = await agent.generate_tz()
            remember(uid, content)
            last_doc[uid] = {"content": content, "type": "tz", "name": agent.tender_name}
            path = await generate_tz_docx(content, agent.tender_name)
            await send_file(context, chat_id, path, "TZ_" + agent.tender_name[:30] + ".docx",
                            "📄 Техническое задание готово!")

        if agent.doc_type in ("criteria_only", "both"):
            await _gen_criteria(uid, chat_id, context)
            return

        if agent.doc_type == "tz_only":
            await context.bot.send_message(
                chat_id=chat_id, text="Нужны критерии допуска?",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Да", callback_data="yes_criteria")],
                    [InlineKeyboardButton("Нет, всё готово", callback_data="no_criteria")]]))
            return
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Generate: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка генерации. /new и попробуй заново.")

async def _gen_criteria(uid, chat_id, context):
    from docx_generator import generate_criteria_docx
    from examples_loader import load_examples
    agent = sessions.get(uid)
    examples_texts = load_examples("criteria")
    examples_block = ""
    if examples_texts:
        examples_block = "ПРИМЕРЫ КРИТЕРИЕВ ИЗ РЕАЛЬНЫХ ТЕНДЕРОВ:\n\n"
        for i, t in enumerate(examples_texts[:5], 1):
            examples_block += "=== Пример " + str(i) + " ===\n" + t[:3000] + "\n\n"
    system = ("Ты эксперт по коммерческим закупкам работ и услуг.\n\n" + examples_block +
              "Составь критерии допуска участников. Используй примеры как образец.\n\n"
              "КРИТИЧЕСКИ ВАЖНО соблюдать формат вывода:\n"
              "- Каждый критерий — три строки подряд:\n"
              "КРИТЕРИЙ: краткое название\n"
              "ТРЕБОВАНИЕ: конкретное измеримое требование\n"
              "ДОКУМЕНТ: что предоставить в подтверждение\n"
              "- Между критериями — пустая строка\n"
              "- НЕ используй markdown (звёздочки, решётки)\n"
              "- НЕ нумеруй критерии (1., 2., 3.)\n"
              "- НЕ добавляй заголовки или вводные фразы\n"
              "- Только сам список критериев\n\n"
              "Составь 6-10 критериев.")
    try:
        await context.bot.send_message(chat_id=chat_id, text="Генерирую критерии допуска...")
        ctx_data = agent._context() if agent else context.user_data.get("custom_ctx", "")
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=3000, system=system,
            messages=[{"role": "user", "content": "Данные закупки:\n" + ctx_data}])
        content = resp.content[0].text
        remember(uid, content)
        name = agent.tender_name if agent else context.user_data.get("custom_direction_name", "Zakupka")
        last_doc[uid] = {"content": content, "type": "criteria", "name": name}
        path = await generate_criteria_docx(content, name)
        await send_file(context, chat_id, path, "Kriterii_" + name[:25] + ".docx",
                        "📋 Критерии допуска готовы!")
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Criteria: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка генерации критериев.")

async def _gen_criteria_custom(uid, chat_id, context, details):
    """Критерии для своего направления — ищет релевантные примеры в базе."""
    from docx_generator import generate_criteria_docx
    from examples_loader import load_examples
    custom_name = context.user_data.get("custom_direction_name", "Закупка")
    examples_texts = load_examples("criteria")
    examples_block = ""
    if examples_texts:
        examples_block = "БАЗА ПРИМЕРОВ КРИТЕРИЕВ (разные направления):\n\n"
        for i, t in enumerate(examples_texts[:8], 1):
            examples_block += "=== Пример " + str(i) + " ===\n" + t[:2500] + "\n\n"
    system = ("Ты эксперт по коммерческим закупкам работ и услуг.\n\n" + examples_block +
              "Направление закупки: " + custom_name + "\n\n"
              "Из базы примеров выше выбери НАИБОЛЕЕ РЕЛЕВАНТНЫЕ данному направлению "
              "и составь критерии допуска по их образцу.\n\n"
              "КРИТИЧЕСКИ ВАЖНО соблюдать формат вывода:\n"
              "- Каждый критерий — три строки подряд:\n"
              "КРИТЕРИЙ: краткое название\n"
              "ТРЕБОВАНИЕ: конкретное измеримое требование\n"
              "ДОКУМЕНТ: что предоставить в подтверждение\n"
              "- Между критериями — пустая строка\n"
              "- НЕ используй markdown (звёздочки, решётки)\n"
              "- НЕ нумеруй критерии\n"
              "- НЕ добавляй заголовки или вводные фразы\n\n"
              "Составь 6-10 критериев.")
    try:
        await context.bot.send_message(chat_id=chat_id, text="Ищу релевантные примеры и генерирую критерии...")
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=3000, system=system,
            messages=[{"role": "user", "content": "Направление: " + custom_name + "\nДетали закупки: " + details}])
        content = resp.content[0].text
        remember(uid, content)
        last_doc[uid] = {"content": content, "type": "criteria", "name": custom_name}
        path = await generate_criteria_docx(content, custom_name)
        await send_file(context, chat_id, path, "Kriterii_" + custom_name[:25] + ".docx",
                        "📋 Критерии допуска готовы!")
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("CriteriaCustom: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. Попробуй ещё раз.")

async def cb_criteria_q(update, context):
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    if q.data == "no_criteria":
        await q.edit_message_text("Хорошо!")
        await send_review(context, q.message.chat_id)
        return
    await q.edit_message_text("Генерирую критерии...")
    await _gen_criteria(uid, q.message.chat_id, context)

# ─── Свободное ТЗ (любое направление) ───────────────────────────────────────
FREE_TZ_SYSTEM = """Ты эксперт по коммерческим закупкам работ и услуг.
Тебе нужно составить Техническое задание для закупки по нестандартному направлению.

Если есть примеры ТЗ по похожим закупкам — используй их как образец структуры.
Если примеров нет — составь полноценное ТЗ самостоятельно: введение, предмет договора, требования к исполнителю, объём работ, сроки, порядок сдачи-приёмки.
Используй таблицы там где уместно (перечень работ, объекты, сроки).
Деловой стиль. Русский язык. Начни с заголовка ТЕХНИЧЕСКОЕ ЗАДАНИЕ."""

FREE_TZ_QUESTIONS_SYSTEM = """Ты эксперт по коммерческим закупкам работ и услуг.
Пользователь хочет составить ТЗ по направлению: {direction}

Задавай конкретные вопросы по одному, чтобы собрать всё необходимое для ТЗ.
Исходя из типа закупки определи какие параметры важны — объём, сроки, требования к исполнителю, адрес объекта, периодичность и т.д.
Задай 4-7 самых важных вопросов. Когда данных достаточно — скажи "Данных достаточно, генерирую ТЗ" и сгенерируй его.
Русский язык."""

async def _free_tz_start(uid, chat_id, context, tz_name):
    """Запускает флоу свободного ТЗ после получения названия."""
    context.user_data["free_tz_name"] = tz_name
    context.user_data["free_tz_docs"] = []
    context.user_data["free_tz_history"] = []
    set_mode(context, "awaiting_free_tz_docs")
    await context.bot.send_message(
        chat_id=chat_id,
        text="Понял! Закупка: " + tz_name + "\n\n"
             "Есть примеры ТЗ по аналогичным закупкам или другие полезные документы? "
             "Загрузи их — использую как образец.\n\n"
             "Когда загрузишь всё нужное — напиши «готово». "
             "Если документов нет — просто напиши «нет» и я задам вопросы.",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Нет документов, задавай вопросы", callback_data="free_tz_no_docs")],
        ])
    )

async def cb_free_tz_nodocs(update, context):
    """Пользователь выбрал 'нет документов' — переходим к вопросам."""
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    tz_name = context.user_data.get("free_tz_name", "закупка")
    await q.edit_message_text("Хорошо, задам вопросы для составления ТЗ.")
    await _free_tz_ask_question(uid, q.message.chat_id, context)

async def _free_tz_ask_question(uid, chat_id, context):
    """Задаёт следующий вопрос по ТЗ через диалог с Claude."""
    tz_name = context.user_data.get("free_tz_name", "закупка")
    history = context.user_data.get("free_tz_history", [])
    docs = context.user_data.get("free_tz_docs", [])
    set_mode(context, "free_tz_qa")

    system = FREE_TZ_QUESTIONS_SYSTEM.replace("{direction}", tz_name)
    docs_ctx = ""
    if docs:
        docs_ctx = "\n\nЗагруженные документы:\n" + "\n\n---\n\n".join(docs[:3])

    messages = [{"role": "user", "content": "Направление закупки: " + tz_name + docs_ctx}]
    for h in history:
        messages.append(h)

    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=1000, system=system, messages=messages)
        reply = resp.content[0].text
        history.append({"role": "assistant", "content": reply})
        context.user_data["free_tz_history"] = history

        # Если Claude говорит что данных достаточно — генерируем ТЗ
        if "данных достаточно" in reply.lower() or "генерирую тз" in reply.lower() or "ТЕХНИЧЕСКОЕ ЗАДАНИЕ" in reply:
            await _free_tz_generate(uid, chat_id, context, from_qa=True, qa_content=reply)
        else:
            await context.bot.send_message(chat_id=chat_id, text=reply)
    except Exception as e:
        logger.error("FreeTZ QA: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. /new и попробуй заново.")

async def _free_tz_generate(uid, chat_id, context, from_qa=False, qa_content=""):
    """Генерирует ТЗ по накопленным данным."""
    from docx_generator import generate_tz_docx
    tz_name = context.user_data.get("free_tz_name", "ТЗ")
    docs = context.user_data.get("free_tz_docs", [])
    history = context.user_data.get("free_tz_history", [])

    if from_qa and "ТЕХНИЧЕСКОЕ ЗАДАНИЕ" in qa_content:
        # Claude уже вернул готовое ТЗ в qa_content
        tz_text = qa_content
    else:
        await context.bot.send_message(chat_id=chat_id, text="Составляю ТЗ...")
        docs_ctx = ""
        if docs:
            docs_ctx = "\n\nПримеры и документы пользователя:\n" + "\n\n---\n\n".join(docs[:4])
        qa_ctx = ""
        if history:
            qa_ctx = "\n\nОтветы пользователя на вопросы:\n" + "\n".join(
                (h["role"].upper() + ": " + h["content"]) for h in history)
        user_msg = ("Направление закупки: " + tz_name + docs_ctx + qa_ctx +
                    "\n\nСоставь полноценное Техническое задание.")
        try:
            resp = claude.messages.create(
                model="claude-sonnet-4-6", max_tokens=8000, system=FREE_TZ_SYSTEM,
                messages=[{"role": "user", "content": user_msg}])
            tz_text = resp.content[0].text
        except Exception as e:
            logger.error("FreeTZ gen: " + str(e), exc_info=True)
            await context.bot.send_message(chat_id=chat_id, text="Ошибка при генерации. Попробуй ещё раз.")
            return

    remember(uid, tz_text)
    last_doc[uid] = {"content": tz_text, "type": "tz", "name": tz_name}
    path = await generate_tz_docx(tz_text, tz_name)
    await send_file(context, chat_id, path, "TZ_" + tz_name[:30] + ".docx", "📄 Техническое задание готово!")
    set_mode(context, None)
    await send_review(context, chat_id)

# ─── Сравнение двух документов ───────────────────────────────────────────────
COMPARE_SYSTEM = """Ты эксперт по коммерческим закупкам и деловым документам.
Тебе дали два документа для сравнения. Задача пользователя указана отдельно.

Проведи детальное сравнение согласно заданию. Структурируй результат:
1. КРАТКИЙ ВЫВОД — 2-3 предложения о главном
2. ОСНОВНЫЕ РАЗЛИЧИЯ — по пунктам, конкретно
3. ПРОТИВОРЕЧИЯ — если есть несоответствия между документами
4. ВЫВОДЫ И РЕКОМЕНДАЦИИ

Если задание конкретное (например "найди различия в объёмах работ") — фокусируйся именно на этом.
Русский язык."""

async def _run_comparison(uid, chat_id, context):
    """Запускает сравнение двух загруженных документов."""
    from docx_generator import generate_tz_docx
    doc1 = context.user_data.get("compare_doc1", "")
    doc2 = context.user_data.get("compare_doc2", "")
    task = context.user_data.get("compare_task", "Найди все различия и противоречия")
    name1 = context.user_data.get("compare_name1", "Документ 1")
    name2 = context.user_data.get("compare_name2", "Документ 2")

    await context.bot.send_message(chat_id=chat_id, text="Сравниваю документы...")
    prompt = ("Документ 1 (" + name1 + "):\n" + doc1[:6000] +
              "\n\n---\n\nДокумент 2 (" + name2 + "):\n" + doc2[:6000] +
              "\n\n---\n\nЗадание пользователя: " + task)
    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=4000, system=COMPARE_SYSTEM,
            messages=[{"role": "user", "content": prompt}])
        result = resp.content[0].text
        remember(uid, result)
        last_doc[uid] = {"content": result, "type": "analysis", "name": "Sravnenie"}
        path = await generate_tz_docx(result, "Sravnenie")
        await send_file(context, chat_id, path, "Sravnenie_dokumentov.docx", "⚖️ Сравнение документов готово!")
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Compare: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка при сравнении. Попробуй ещё раз.")

# ─── Правки (review) ─────────────────────────────────────────────────────────
async def cb_review(update, context):
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    if q.data == "review_ok":
        await q.edit_message_text("Отлично! Обращайся если что. /new для нового запроса.")
        set_mode(context, None)
        return
    set_mode(context, "awaiting_review_edits")
    await q.edit_message_text("Напиши замечания - внесу правки:")

async def apply_review_edits(uid, chat_id, update, context, edits):
    from docx_generator import generate_tz_docx, generate_criteria_docx
    doc_info = last_doc.get(uid, {})
    if not doc_info:
        await context.bot.send_message(chat_id=chat_id, text="Не нашёл документ. /new")
        set_mode(context, None); return
    await context.bot.send_message(chat_id=chat_id, text="Вношу правки...")
    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=6000, system=REVIEW_SYSTEM,
            messages=[{"role": "user", "content": "Документ:\n" + doc_info["content"] +
                       "\n\nПравки:\n" + edits + "\n\nВерни исправленный документ."}])
        new_content = resp.content[0].text
        last_doc[uid]["content"] = new_content
        remember(uid, new_content)
        doc_type = doc_info.get("type", "tz"); name = doc_info.get("name", "Doc")
        if doc_type == "criteria":
            path = await generate_criteria_docx(new_content, name)
            await send_file(context, chat_id, path, "Kriterii_" + name[:25] + ".docx", "📋 Критерии обновлены!")
        else:
            path = await generate_tz_docx(new_content, name)
            fname_map = {"letter": "Pismo.docx", "negotiation": "Scenariy.docx",
                         "analysis": "Ekspertiza.docx", "edited": "Dokument.docx"}
            cap_map = {"letter": "✉️ Письмо обновлено!", "negotiation": "🤝 Сценарий обновлён!",
                       "analysis": "🔎 Заключение обновлено!", "edited": "📄 Документ обновлён!"}
            fname = fname_map.get(doc_type, "TZ_" + name[:25] + ".docx")
            cap = cap_map.get(doc_type, "📄 ТЗ обновлено!")
            await send_file(context, chat_id, path, fname, cap)
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Edit: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. Попробуй ещё раз.")
        set_mode(context, "awaiting_review_edits")

# ─── Переговоры ──────────────────────────────────────────────────────────────
async def cb_neg(update, context):
    q = update.callback_query; await q.answer()
    parts = q.data.split("_"); step = int(parts[1]); choice = parts[2]
    if choice == "custom":
        context.user_data["neg_custom_step"] = step
        set_mode(context, "negotiation")
        await q.edit_message_text(NEGOTIATION_STEPS[step]["q"] + "\n\nВведи свой вариант:")
        return
    answer = NEGOTIATION_STEPS[step]["opts"][int(choice)]
    await q.edit_message_text(NEGOTIATION_STEPS[step]["q"] + "\n> " + answer)
    await _save_neg(q.message.chat_id, context, step, answer)

async def _save_neg(chat_id, context, step, answer):
    answers = context.user_data.setdefault("neg_answers", {})
    answers[step] = {"q": NEGOTIATION_STEPS[step]["q"], "a": answer}
    next_step = step + 1
    context.user_data["neg_step"] = next_step
    if next_step < len(NEGOTIATION_STEPS):
        await context.bot.send_message(chat_id=chat_id, text=NEGOTIATION_STEPS[next_step]["q"],
                                        reply_markup=neg_kb(next_step))
        return
    await context.bot.send_message(chat_id=chat_id, text="Составляю сценарий переговоров...")
    await _gen_negotiation(chat_id, context)

async def _gen_negotiation(chat_id, context):
    answers = context.user_data.get("neg_answers", {})
    ctx = "\n".join(answers[i]["q"] + ": " + answers[i]["a"]
                    for i in range(len(NEGOTIATION_STEPS)) if i in answers)
    try:
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=4000, system=NEGOTIATION_SYSTEM,
            messages=[{"role": "user", "content": "Данные:\n" + ctx + "\n\nСоставь конкретный сценарий без воды."}])
        content = resp.content[0].text
        uid = chat_id
        remember(uid, content)
        last_doc[uid] = {"content": content, "type": "negotiation", "name": "Scenariy"}
        set_mode(context, None)
        await context.bot.send_message(chat_id=chat_id, text="Сценарий готов! В каком формате сохранить?",
                                        reply_markup=word_pdf_kb())
    except Exception as e:
        logger.error("Neg: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. /new и попробуй заново.")

async def cb_fmt(update, context):
    from docx_generator import generate_tz_docx, generate_pdf
    q = update.callback_query; await q.answer()
    uid = update.effective_user.id
    doc_info = last_doc.get(uid, {})
    if not doc_info:
        await q.edit_message_text("Не нашёл документ."); return
    fmt = q.data.replace("fmt_", "")
    content = doc_info["content"]; name = doc_info.get("name", "Doc")
    base = "Scenariy" if doc_info.get("type") == "negotiation" else "Dokument"
    await q.edit_message_text("Создаю файл(ы)...")
    chat_id = q.message.chat_id
    try:
        if fmt in ("docx", "both"):
            path = await generate_tz_docx(content, name)
            await send_file(context, chat_id, path, base + ".docx", "📄 Word готов!")
        if fmt in ("pdf", "both"):
            path = await generate_pdf(content, name)
            await send_file(context, chat_id, path, base + ".pdf", "📕 PDF готов!")
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Fmt: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка при создании файла.")

# ─── Письма ──────────────────────────────────────────────────────────────────
async def cb_letter_type(update, context):
    q = update.callback_query; await q.answer()
    if q.data == "letter_reply":
        context.user_data["letter_mode"] = "reply"
        set_mode(context, "awaiting_letter_original")
        await q.edit_message_text(
            "Отправь оригинал письма: файл (Word/PDF/txt), фото или вставь текст:")
    else:
        context.user_data["letter_mode"] = "new"
        set_mode(context, "awaiting_letter_input")
        await q.edit_message_text(
            "Есть черновик письма? Загрузи файл/фото или вставь текст - доработаю.\n\n"
            "Если черновика нет - просто напиши кому, о чём и что сообщить:")

async def _gen_letter(uid, chat_id, context, comment):
    from docx_generator import generate_tz_docx
    original = context.user_data.get("letter_original", "")
    mode = context.user_data.get("letter_mode", "new")
    await context.bot.send_message(chat_id=chat_id, text="Составляю письмо...")
    try:
        if mode == "reply" and original:
            prompt = ("Оригинальное письмо:\n" + original +
                      "\n\nКомментарий (как ответить): " + comment +
                      "\n\nСоставь профессиональный ответ на это письмо. "
                      "Если комментарий просит понять из контекста - сам реши как корректно ответить.")
            fname = "Otvet_na_pismo.docx"; cap = "✉️ Ответное письмо готово!"
        elif mode == "edit_draft" and original:
            prompt = ("Черновик письма:\n" + original +
                      "\n\nКомментарий по доработке: " + comment +
                      "\n\nДоработай письмо. Верни полный текст готового письма.")
            fname = "Pismo.docx"; cap = "✉️ Письмо готово!"
        else:
            prompt = "Задание: " + comment + "\n\nСоставь профессиональное деловое письмо."
            fname = "Pismo.docx"; cap = "✉️ Письмо готово!"
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000, system=LETTER_SYSTEM,
            messages=[{"role": "user", "content": prompt}])
        content = resp.content[0].text
        remember(uid, content)
        last_doc[uid] = {"content": content, "type": "letter", "name": "Pismo"}
        path = await generate_tz_docx(content, "Pismo")
        await send_file(context, chat_id, path, fname, cap)
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Letter: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. Попробуй ещё раз.")

# ─── Анализ ──────────────────────────────────────────────────────────────────
async def _run_analysis(uid, chat_id, context, comment):
    from docx_generator import generate_tz_docx
    doc_text = context.user_data.get("analysis_doc", "")
    if not doc_text:
        await context.bot.send_message(chat_id=chat_id, text="Документ не найден. Загрузи заново.")
        set_mode(context, "awaiting_analysis_doc"); return
    await context.bot.send_message(chat_id=chat_id, text="Провожу экспертизу, это займёт около минуты...")
    try:
        focus = "" if comment.lower().strip() in ("анализируй", "анализ", "") else "\nОсобое внимание: " + comment
        prompt = ("Проведи экспертизу документа организатора коммерческого тендера." + focus +
                  "\n\nДОКУМЕНТ:\n" + doc_text[:12000] +
                  "\n\nВыдай полное структурированное заключение:\n"
                  "ЭКСПЕРТНОЕ ЗАКЛЮЧЕНИЕ\n"
                  "1. КРАТКОЕ РЕЗЮМЕ\n2. ОШИБКИ И ПРОТИВОРЕЧИЯ\n"
                  "3. ОГРАНИЧИВАЮЩИЕ ТРЕБОВАНИЯ\n4. НЕЧЁТКИЕ ФОРМУЛИРОВКИ\n"
                  "5. НЕДОСТАЮЩИЕ ТРЕБОВАНИЯ\n6. КОНКРЕТНЫЕ ПРЕДЛОЖЕНИЯ ПО ДОРАБОТКЕ\n"
                  "По каждому замечанию давай конкретное предложение с готовой формулировкой.")
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=8000, system=ANALYSIS_SYSTEM,
            messages=[{"role": "user", "content": prompt}])
        analysis = resp.content[0].text
        remember(uid, analysis)
        last_doc[uid] = {"content": analysis, "type": "analysis", "name": "Ekspertiza"}
        path = await generate_tz_docx(analysis, "Ekspertiza")
        await send_file(context, chat_id, path, "Ekspertiza.docx", "🔎 Экспертное заключение готово!")
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("Analysis: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка при анализе. Попробуй ещё раз.")

# ─── Документ в чате ─────────────────────────────────────────────────────────
async def cb_doc_action(update, context):
    q = update.callback_query; await q.answer()
    action = q.data.replace("docact_", "")
    context.user_data["doc_action"] = action
    set_mode(context, "awaiting_doc_comment")
    prompts = {
        "analyze": "Напиши на что обратить внимание при анализе (или просто 'анализируй'):",
        "edit": "Напиши комментарий - что изменить, добавить или убрать:",
        "reply": "Напиши как ответить и что включить (или 'пойми из контекста сам'):",
    }
    await q.edit_message_text(prompts.get(action, "Напиши комментарий:"))

async def _run_doc_action(uid, chat_id, context, comment):
    from docx_generator import generate_tz_docx
    doc_text = context.user_data.get("received_doc_text", "")
    action = context.user_data.get("doc_action", "analyze")
    if not doc_text:
        await context.bot.send_message(chat_id=chat_id, text="Документ не найден. Загрузи снова.")
        set_mode(context, None); return

    if action == "analyze":
        context.user_data["analysis_doc"] = doc_text
        await _run_analysis(uid, chat_id, context, comment)
        return

    await context.bot.send_message(chat_id=chat_id, text="Работаю...")
    try:
        if action == "edit":
            system = REVIEW_SYSTEM
            prompt = ("Документ:\n" + doc_text + "\n\nКомментарий: " + comment +
                      "\n\nВнеси правки и верни полный исправленный текст.")
            fname = "Otredaktirovan.docx"; cap = "✏️ Отредактированный документ готов!"
            dtype = "edited"
        else:  # reply
            system = LETTER_SYSTEM
            prompt = ("Входящее письмо:\n" + doc_text + "\n\nКомментарий (как ответить): " + comment +
                      "\n\nСоставь профессиональный ответ. Если комментарий просит понять "
                      "из контекста - сам реши как корректно ответить.")
            fname = "Otvet_na_pismo.docx"; cap = "✉️ Ответное письмо готово!"
            dtype = "letter"
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=4000, system=system,
            messages=[{"role": "user", "content": prompt}])
        content = resp.content[0].text
        remember(uid, content)
        last_doc[uid] = {"content": content, "type": dtype, "name": "Dokument"}
        path = await generate_tz_docx(content, "Dokument")
        await send_file(context, chat_id, path, fname, cap)
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("DocAction: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. Попробуй ещё раз.")

# ─── Фото письма ─────────────────────────────────────────────────────────────
async def cb_photo_action(update, context):
    q = update.callback_query; await q.answer()
    action = q.data.replace("photo_act_", "")
    context.user_data["photo_action"] = action
    set_mode(context, "awaiting_photo_comment")
    prompts = {
        "reply": "Напиши как ответить и что включить (или 'пойми из контекста сам'):",
        "edit": "Напиши комментарий - что изменить, добавить или убрать:",
        "analyze": "Напиши на что обратить внимание (или просто 'анализируй'):",
    }
    await q.edit_message_text(prompts.get(action, "Напиши комментарий:"))

async def _run_photo_action(uid, chat_id, context, comment):
    from docx_generator import generate_tz_docx
    original = context.user_data.get("last_photo_text", "")
    action = context.user_data.get("photo_action", "reply")
    if not original:
        await context.bot.send_message(chat_id=chat_id, text="Текст письма не найден. Отправь фото заново.")
        set_mode(context, None); return

    if action == "analyze":
        context.user_data["analysis_doc"] = original
        await _run_analysis(uid, chat_id, context, comment)
        return

    await context.bot.send_message(chat_id=chat_id, text="Работаю...")
    try:
        if action == "reply":
            system = LETTER_SYSTEM
            prompt = ("Входящее письмо:\n" + original + "\n\nКомментарий (как ответить): " + comment +
                      "\n\nСоставь профессиональный ответ. Если просят понять из контекста - реши сам.")
            fname = "Otvet_na_pismo.docx"; cap = "✉️ Ответное письмо готово!"
        else:  # edit
            system = REVIEW_SYSTEM
            prompt = ("Письмо:\n" + original + "\n\nКомментарий: " + comment +
                      "\n\nОтредактируй письмо. Верни полный исправленный текст.")
            fname = "Otredaktirovannoe_pismo.docx"; cap = "✏️ Отредактированное письмо готово!"
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000, system=system,
            messages=[{"role": "user", "content": prompt}])
        content = resp.content[0].text
        remember(uid, content)
        last_doc[uid] = {"content": content, "type": "letter", "name": "Pismo"}
        path = await generate_tz_docx(content, "Pismo")
        await send_file(context, chat_id, path, fname, cap)
        set_mode(context, None)
        await send_review(context, chat_id)
    except Exception as e:
        logger.error("PhotoAction: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка. Попробуй ещё раз.")

# ─── Сохранить из чата ───────────────────────────────────────────────────────
async def save_last(uid, chat_id, context):
    from docx_generator import generate_tz_docx
    msgs_list = bot_msgs.get(uid, [])
    content = msgs_list[-1] if msgs_list else None
    if not content:
        history = context.user_data.get("chat_history", [])
        for m in reversed(history):
            if m["role"] == "assistant" and isinstance(m["content"], str):
                content = m["content"]; break
    if not content:
        await context.bot.send_message(chat_id=chat_id, text="Нечего сохранять."); return
    await context.bot.send_message(chat_id=chat_id, text="Сохраняю в Word...")
    try:
        path = await generate_tz_docx(content, "Dokument")
        await send_file(context, chat_id, path, "Dokument.docx", "📄 Сохранено в Word!")
    except Exception as e:
        logger.error("Save: " + str(e), exc_info=True)
        await context.bot.send_message(chat_id=chat_id, text="Ошибка при сохранении.")

async def cb_save_to_word(update, context):
    q = update.callback_query; await q.answer()
    await save_last(update.effective_user.id, q.message.chat_id, context)

# ─── Голосовой умный режим ───────────────────────────────────────────────────
async def handle_voice_smart(update, context, transcribed):
    uid = update.effective_user.id
    chat_id = update.message.chat_id
    try:
        import re as re_mod
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=500,
            system=VOICE_EXTRACT_SYSTEM,
            messages=[{"role": "user", "content": transcribed}])
        raw = resp.content[0].text.strip()
        raw = re_mod.sub(r"```json|```", "", raw).strip()
        data = json.loads(raw)
    except Exception as e:
        logger.error("Voice extract: " + str(e))
        return False

    intent = data.get("intent", "chat")
    direction = data.get("direction")
    params = data.get("params", {})
    task = data.get("original_task", transcribed)
    if intent == "chat":
        return False

    sessions.pop(uid, None); last_doc.pop(uid, None); context.user_data.clear()

    if intent == "analysis":
        set_mode(context, "awaiting_analysis_doc")
        await update.message.reply_text("Загрузи документ для экспертизы (Word, PDF, txt или фото).")
        return True

    if intent == "letter":
        context.user_data["letter_mode"] = "new"
        context.user_data["letter_original"] = ""
        await _gen_letter(uid, chat_id, context, task)
        return True

    if intent == "negotiation":
        neg_answers = {}
        param_map = {
            0: ["что закупаем", "закупаем", "услуга", "работа", "предмет"],
            1: ["кто придёт", "представитель", "директор"],
            2: ["нмц", "цена", "стоимость", "бюджет"],
            3: ["снижение", "снизить", "скидка", "процент"],
            4: ["конкурент", "альтернатив", "участник"],
            5: ["цел", "дополнительно"],
        }
        for step_idx, keywords in param_map.items():
            for key, val in params.items():
                if any(kw in key.lower() for kw in keywords):
                    neg_answers[step_idx] = {"q": NEGOTIATION_STEPS[step_idx]["q"], "a": str(val)}
                    break
        context.user_data["neg_answers"] = neg_answers
        set_mode(context, "negotiation")
        next_step = next((i for i in range(len(NEGOTIATION_STEPS)) if i not in neg_answers), None)
        if next_step is None:
            context.user_data["neg_step"] = len(NEGOTIATION_STEPS)
            await update.message.reply_text("Все данные есть! Составляю сценарий...")
            await _gen_negotiation(chat_id, context)
        else:
            context.user_data["neg_step"] = next_step
            filled = len(neg_answers)
            prefix = ("Заполнил " + str(filled) + " из " + str(len(NEGOTIATION_STEPS)) +
                      " параметров из голосового.\n\n") if filled else "Запускаю переговоры!\n\n"
            await update.message.reply_text(prefix + NEGOTIATION_STEPS[next_step]["q"],
                                             reply_markup=neg_kb(next_step))
        return True

    if intent in ("tz", "criteria", "both"):
        doc_type_map = {"tz": "tz_only", "criteria": "criteria_only", "both": "both"}
        doc_type = doc_type_map.get(intent, "tz_only")
        context.user_data["doc_type"] = doc_type
        if not direction:
            await update.message.reply_text("Понял задачу: " + task + "\nВыбери направление:",
                                             reply_markup=dir_kb(doc_type))
            return True
        context.user_data["direction"] = direction
        agent = TenderAgent(direction=direction, doc_type=doc_type)
        sessions[uid] = agent
        dir_questions = QUESTIONS.get(direction, [])
        for q_item in dir_questions:
            q_text = q_item["question"].lower()
            for pk, pv in params.items():
                if any(word in q_text for word in pk.lower().split() if len(word) > 3):
                    agent.answers.append({"question": q_item["question"], "answer": str(pv)})
                    agent.current_q += 1
                    break
        filled = len(agent.answers)
        if filled > 0:
            ft = "\n".join("- " + a["question"] + ": " + a["answer"] for a in agent.answers)
            await update.message.reply_text("Заполнил из голосового " + str(filled) + " параметров:\n" + ft)
        set_mode(context, "answering")
        result = await agent.get_next_question()
        if result["status"] == "question":
            await send_q(update.message, result)
        else:
            await update.message.reply_text("Все данные есть! Генерирую...")
            await do_generate(uid, chat_id, context)
        return True
    return False

# ─── Роутер документов ───────────────────────────────────────────────────────
async def route_document(update, context):
    uid = update.effective_user.id
    chat_id = update.message.chat_id
    mode = get_mode(context)
    doc = update.message.document
    mime = doc.mime_type or ""
    fname = doc.file_name or ""

    if mime.startswith("image/"):
        await route_photo(update, context); return

    await update.message.reply_text("Читаю документ...")
    doc_text = await extract_doc_text(update)
    if not doc_text:
        await update.message.reply_text(
            "Не смог прочитать файл '" + fname + "'.\n"
            "Поддерживаю: .docx, .pdf, .txt, .xlsx (старый .doc и .xls — нет).\n"
            "Если файл .docx/.xlsx и не читается — пересохрани его заново.")
        return

    if mode == "awaiting_customer_doc":
        await _process_customer_doc(uid, chat_id, update, context, doc_text)
        return

    if mode in ("awaiting_letter_original", "awaiting_letter_input"):
        context.user_data["letter_original"] = doc_text
        if mode == "awaiting_letter_input":
            context.user_data["letter_mode"] = "edit_draft"
            set_mode(context, "awaiting_letter_task")
            await update.message.reply_text("Прочитал черновик! Что доработать, изменить, добавить?")
        else:
            context.user_data["letter_mode"] = "reply"
            set_mode(context, "awaiting_letter_task")
            await update.message.reply_text(
                "Прочитал письмо! Как ответить, что включить? (или 'пойми из контекста сам')")
        return

    if mode == "awaiting_analysis_doc":
        context.user_data["analysis_doc"] = doc_text
        set_mode(context, "awaiting_analysis_comment")
        await update.message.reply_text(
            "Документ получен. На что обратить особое внимание? (или просто 'анализируй')")
        return

    # ─── Новые режимы ───
    if mode == "awaiting_free_tz_docs":
        docs = context.user_data.setdefault("free_tz_docs", [])
        docs.append(fname + ":\n" + doc_text[:4000])
        await update.message.reply_text(
            "Получил: " + fname + " ✓\n\n"
            "Загрузи ещё файлы если нужно, или напиши «готово» чтобы я сформировал ТЗ.")
        return

    if mode == "awaiting_compare_doc1":
        context.user_data["compare_doc1"] = doc_text
        context.user_data["compare_name1"] = fname
        set_mode(context, "awaiting_compare_doc2")
        await update.message.reply_text(
            "Получил: " + fname + " ✓\n\nТеперь загрузи второй документ:")
        return

    if mode == "awaiting_compare_doc2":
        context.user_data["compare_doc2"] = doc_text
        context.user_data["compare_name2"] = fname
        set_mode(context, "awaiting_compare_task")
        await update.message.reply_text(
            "Получил: " + fname + " ✓\n\n"
            "Что именно сравнивать? Напиши конкретную задачу\n"
            "(например: «найди различия в объёмах работ», «проверь нет ли противоречий», «сравни требования к исполнителю»):")
        return

    # Вне режимов — спрашиваем что делать
    context.user_data["received_doc_text"] = doc_text
    context.user_data["received_doc_name"] = fname
    remember(uid, doc_text[:500])
    await update.message.reply_text(
        "Получил: " + fname + "\n\nЧто с ним делаем?",
        reply_markup=doc_action_kb())

async def _process_customer_doc(uid, chat_id, update, context, doc_text):
    await update.message.reply_text("Анализирую что уже есть...")
    direction = context.user_data.get("direction", "cleaning")
    q_list = "\n".join("- " + q["question"] for q in QUESTIONS.get(direction, []))
    try:
        import re as re_mod
        resp = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000,
            system='Проанализируй документ. Верни JSON: {"found": {"вопрос": "ответ"}, "missing": ["вопрос"]}. Только JSON.',
            messages=[{"role": "user", "content": "Вопросы:\n" + q_list + "\n\nДокумент:\n" + doc_text[:6000]}])
        raw = re_mod.sub(r"```json|```", "", resp.content[0].text).strip()
        analysis = json.loads(raw)
    except:
        analysis = {"found": {}, "missing": [q["question"] for q in QUESTIONS.get(direction, [])]}
    found = analysis.get("found", {}); missing = analysis.get("missing", [])
    context.user_data["prefilled"] = found
    if found:
        await update.message.reply_text(
            "Нашёл в документе:\n" + "\n".join("- " + k + ": " + str(v) for k, v in found.items()) +
            "\n\nОсталось уточнить: " + str(len(missing)) + " вопр.")
    doc_type = context.user_data.get("doc_type", "tz_only")
    all_q = QUESTIONS.get(direction, [])
    filtered = [q for q in all_q if any(m.lower() in q["question"].lower() for m in missing)] if missing else []
    agent = TenderAgent(direction=direction, doc_type=doc_type)
    if filtered: agent.questions = filtered
    for fq, fa in found.items():
        agent.answers.append({"question": fq, "answer": str(fa), "prefilled": True})
    sessions[uid] = agent
    set_mode(context, "answering")
    if not missing:
        await update.message.reply_text("Все данные есть! Генерирую...")
        await do_generate(uid, chat_id, context)
        return
    result = await agent.get_next_question()
    await send_q(update.message, result)

# ─── Роутер фото ─────────────────────────────────────────────────────────────
async def route_photo(update, context):
    uid = update.effective_user.id
    mode = get_mode(context)
    caption = update.message.caption or ""

    result = await get_image_b64(update)
    if not result:
        await update.message.reply_text("Не смог получить фото."); return
    b64, mime = result

    # В режиме письма/анализа — фото это оригинал документа
    if mode in ("awaiting_letter_original", "awaiting_letter_input", "awaiting_analysis_doc"):
        await update.message.reply_text("Читаю текст с фото...")
        try:
            resp = claude.messages.create(
                model="claude-sonnet-4-6", max_tokens=3000,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                    {"type": "text", "text": "Распознай весь текст с изображения. Выведи только текст."}]}])
            text = resp.content[0].text
        except Exception as e:
            logger.error("Photo OCR: " + str(e))
            await update.message.reply_text("Не смог распознать текст."); return
        if mode == "awaiting_analysis_doc":
            context.user_data["analysis_doc"] = text
            set_mode(context, "awaiting_analysis_comment")
            await update.message.reply_text("Текст распознан. На что обратить внимание? (или 'анализируй')")
        elif mode == "awaiting_letter_input":
            context.user_data["letter_original"] = text
            context.user_data["letter_mode"] = "edit_draft"
            set_mode(context, "awaiting_letter_task")
            await update.message.reply_text("Прочитал черновик! Что доработать?")
        else:
            context.user_data["letter_original"] = text
            context.user_data["letter_mode"] = "reply"
            set_mode(context, "awaiting_letter_task")
            await update.message.reply_text("Прочитал письмо! Как ответить? (или 'пойми из контекста')")
        return

    # Обычный режим — анализ фото + поиск
    await update.message.reply_text("Смотрю на фото...")
    try:
        vision_prompt = ("Проанализируй изображение:\n"
                         "1. Опиши что изображено\n"
                         "2. Если есть текст - распознай полностью\n"
                         "3. Составь поисковый запрос\n")
        if caption: vision_prompt += "Вопрос пользователя: " + caption + "\n"
        vision_prompt += "\nФормат:\nОПИСАНИЕ: ...\nТЕКСТ: ...\nПОИСК: ..."
        vision = claude.messages.create(
            model="claude-sonnet-4-6", max_tokens=1500,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                {"type": "text", "text": vision_prompt}]}])
        vtext = vision.content[0].text
        desc = ocr = ""; search_q = caption or ""
        for line in vtext.split("\n"):
            line = line.strip()
            if line.startswith("ОПИСАНИЕ:"): desc = line.replace("ОПИСАНИЕ:", "").strip()
            elif line.startswith("ТЕКСТ:"): ocr = line.replace("ТЕКСТ:", "").strip()
            elif line.startswith("ПОИСК:"): search_q = line.replace("ПОИСК:", "").strip()
        if not desc and not ocr:
            desc = vtext.strip(); search_q = caption or desc[:80]

        context.user_data["last_photo_desc"] = desc
        context.user_data["last_photo_text"] = ocr if ocr and ocr != "текста нет" else desc

        search_text = ""
        if search_q:
            try:
                sr = claude.messages.create(model="claude-sonnet-4-6", max_tokens=1500,
                    tools=[WEB_SEARCH_TOOL],
                    messages=[{"role": "user", "content": "Найди информацию: " + search_q}])
                sm = [{"role": "user", "content": "Найди: " + search_q}]
                while sr.stop_reason == "tool_use":
                    tr = [{"type": "tool_result", "tool_use_id": b.id, "content": b.input.get("query", "")}
                          for b in sr.content if b.type == "tool_use"]
                    sm.append({"role": "assistant", "content": sr.content})
                    sm.append({"role": "user", "content": tr})
                    sr = claude.messages.create(model="claude-sonnet-4-6", max_tokens=1500,
                        tools=[WEB_SEARCH_TOOL], messages=sm)
                search_text = "".join(b.text for b in sr.content if hasattr(b, "text"))
            except Exception as se:
                logger.error("Search: " + str(se))

        parts = []
        if desc: parts.append("На фото: " + desc)
        if ocr and ocr != "текста нет": parts.append("\nТекст с фото:\n" + ocr)
        if search_text: parts.append("\nНашёл в интернете:\n" + search_text)
        final = "\n".join(parts) or "Не смог описать фото."
        if len(final) > 4000: final = final[:4000] + "..."
        remember(uid, final)
        await update.message.reply_text(final)

        all_text = (desc + " " + ocr).lower()
        if any(w in all_text for w in ["уважаем", "прошу", "сообщаем", "исх.", "вх.", "направляем", "настоящим"]):
            await update.message.reply_text(
                "Похоже на официальное письмо. Что с ним сделать?",
                reply_markup=photo_action_kb())
    except Exception as e:
        logger.error("Photo: " + str(e), exc_info=True)
        await update.message.reply_text("Не смог обработать фото.")

# ─── Главный текстовый роутер ────────────────────────────────────────────────
async def chat_handler(update, context):
    uid = update.effective_user.id
    if not is_allowed(uid):
        await update.message.reply_text("Нет доступа."); return
    save_user(uid, update.effective_user.username or "")
    chat_id = update.message.chat_id

    # Документ
    if update.message.document:
        await route_document(update, context); return
    # Фото
    if update.message.photo:
        await route_photo(update, context); return

    mode = get_mode(context)

    # Голос: вне режимов — умная обработка; в режиме — как текст
    if update.message.voice:
        file = await update.message.voice.get_file()
        data = await file.download_as_bytearray()
        try:
            text = await transcribe_voice(bytes(data))
        except Exception as e:
            logger.error("Voice: " + str(e))
            await update.message.reply_text("Не смог распознать голос."); return
        await update.message.reply_text("Услышал: " + text[:150])
        if mode is None:
            handled = await handle_voice_smart(update, context, text)
            if handled: return
    else:
        text = (update.message.text or "").strip()

    if not text: return

    # Навигационная клавиатура
    if "/new" in text:
        sessions.pop(uid, None); last_doc.pop(uid, None); context.user_data.clear()
        await update.message.reply_text("Что делаем?", reply_markup=menu_kb()); return
    if "/help" in text:
        await cmd_help(update, context); return

    # ── Роутинг по режимам ──
    if mode == "answering":
        agent = sessions.get(uid)
        if not agent:
            set_mode(context, None)
            await update.message.reply_text("Сессия устарела. /new"); return
        await _handle_answer(uid, chat_id, update, context, text); return

    if mode == "negotiation":
        step = context.user_data.pop("neg_custom_step", context.user_data.get("neg_step", 0))
        await _save_neg(chat_id, context, step, text); return

    if mode == "awaiting_review_edits":
        set_mode(context, None)
        await apply_review_edits(uid, chat_id, update, context, text); return

    if mode == "awaiting_custom_direction":
        context.user_data["custom_direction_name"] = text
        set_mode(context, "awaiting_custom_details")
        await update.message.reply_text(
            "Принял: " + text + "\nОпиши кратко закупку - объём, сроки, особенности:")
        return

    if mode == "awaiting_custom_details":
        context.user_data["custom_ctx"] = text
        await _gen_criteria_custom(uid, chat_id, context, text); return

    if mode == "awaiting_letter_original":
        context.user_data["letter_original"] = text
        context.user_data["letter_mode"] = "reply"
        set_mode(context, "awaiting_letter_task")
        await update.message.reply_text("Прочитал! Как ответить, что включить? (или 'пойми из контекста')")
        return

    if mode == "awaiting_letter_input":
        # Текст = задание на новое письмо (черновика нет)
        context.user_data["letter_mode"] = "new"
        set_mode(context, None)
        await _gen_letter(uid, chat_id, context, text); return

    if mode == "awaiting_letter_task":
        set_mode(context, None)
        await _gen_letter(uid, chat_id, context, text); return

    if mode == "awaiting_analysis_doc":
        if len(text) > 50:
            context.user_data["analysis_doc"] = text
            set_mode(context, "awaiting_analysis_comment")
            await update.message.reply_text("Текст получен. На что обратить внимание? (или 'анализируй')")
        else:
            await update.message.reply_text("Загрузи документ или вставь текст ТЗ (минимум пару абзацев).")
        return

    if mode == "awaiting_analysis_comment":
        set_mode(context, None)
        await _run_analysis(uid, chat_id, context, text); return

    if mode == "awaiting_doc_comment":
        set_mode(context, None)
        await _run_doc_action(uid, chat_id, context, text); return

    if mode == "awaiting_photo_comment":
        set_mode(context, None)
        await _run_photo_action(uid, chat_id, context, text); return

    # ── Новые режимы ──
    if mode == "awaiting_free_tz_name":
        set_mode(context, None)
        await _free_tz_start(uid, chat_id, context, text); return

    if mode == "awaiting_free_tz_docs":
        tl_lower = text.lower().strip()
        if tl_lower in ("готово", "готов", "всё", "всё загрузил", "загрузил", "нет", "нет документов", "без документов"):
            docs = context.user_data.get("free_tz_docs", [])
            if docs:
                await update.message.reply_text("Отлично, документы получены! Генерирую ТЗ...")
                await _free_tz_generate(uid, chat_id, context)
            else:
                await update.message.reply_text("Хорошо, задам вопросы для составления ТЗ.")
                await _free_tz_ask_question(uid, chat_id, context)
        else:
            await update.message.reply_text("Загружай файлы, или напиши «готово» чтобы я начал составлять ТЗ, или «нет» если документов нет.")
        return

    if mode == "free_tz_qa":
        history = context.user_data.get("free_tz_history", [])
        history.append({"role": "user", "content": text})
        context.user_data["free_tz_history"] = history
        await _free_tz_ask_question(uid, chat_id, context); return

    if mode == "awaiting_compare_task":
        context.user_data["compare_task"] = text
        set_mode(context, None)
        await _run_comparison(uid, chat_id, context); return

    # ── Вне режимов ──
    tl = text.lower()
    if any(w in tl for w in ["сохрани", "в ворд", "в word", "сделай файл"]):
        await save_last(uid, chat_id, context); return

    intent = detect_intent(text)
    if intent:
        sessions.pop(uid, None); last_doc.pop(uid, None); context.user_data.clear()
        await _handle_intent_text(update, context, intent); return

    # Обычный чат
    history = context.user_data.get("chat_history", [])
    photo_ctx = context.user_data.get("last_photo_desc", "")
    system = CHAT_SYSTEM + ("\n\nКонтекст фото: " + photo_ctx if photo_ctx else "")
    history.append({"role": "user", "content": text})
    if len(history) > 20: history = history[-20:]
    await update.message.reply_text("Думаю...")
    try:
        resp = claude.messages.create(model="claude-sonnet-4-6", max_tokens=2000,
                                       system=system, tools=[WEB_SEARCH_TOOL], messages=history)
        msgs = list(history)
        while resp.stop_reason == "tool_use":
            tr = [{"type": "tool_result", "tool_use_id": b.id, "content": b.input.get("query", "")}
                  for b in resp.content if b.type == "tool_use"]
            msgs.append({"role": "assistant", "content": resp.content})
            msgs.append({"role": "user", "content": tr})
            resp = claude.messages.create(model="claude-sonnet-4-6", max_tokens=2000,
                                           system=system, tools=[WEB_SEARCH_TOOL], messages=msgs)
        reply = "".join(b.text for b in resp.content if hasattr(b, "text")) or "Не знаю что ответить."
        remember(uid, reply)
        history.append({"role": "assistant", "content": reply})
        context.user_data["chat_history"] = history
        if len(reply) > 4000: reply = reply[:4000] + "..."
        await update.message.reply_text(reply)
        if len(reply) > 500:
            await update.message.reply_text("Сохранить в Word?",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("📄 Сохранить в Word", callback_data="save_to_word")]]))
    except Exception as e:
        logger.error("Chat: " + str(e), exc_info=True)
        await update.message.reply_text("Что-то сломалось. Попробуй ещё раз.")

async def _handle_intent_text(update, context, intent):
    uid = update.effective_user.id
    chat_id = update.message.chat_id
    if intent == "menu_negotiation":
        context.user_data["neg_answers"] = {}; context.user_data["neg_step"] = 0
        set_mode(context, "negotiation")
        await update.message.reply_text("Запускаю сценарий переговоров!\n\n" + NEGOTIATION_STEPS[0]["q"],
                                         reply_markup=neg_kb(0))
    elif intent == "menu_analysis":
        set_mode(context, "awaiting_analysis_doc")
        await update.message.reply_text(
            "Загрузи документ для экспертизы (Word, PDF, txt, фото или вставь текст).")
    elif intent == "menu_letter":
        await update.message.reply_text("Что нужно?", reply_markup=letter_type_kb())
    elif intent == "menu_compare":
        set_mode(context, "awaiting_compare_doc1")
        await update.message.reply_text("Загрузи первый документ (Word, PDF, Excel, txt):")
    else:
        doc_map = {"menu_tz": "tz_only", "menu_criteria": "criteria_only"}
        doc_type = doc_map.get(intent, "tz_only")
        context.user_data["doc_type"] = doc_type
        await update.message.reply_text("Выбери направление закупки:", reply_markup=dir_kb(doc_type))

# ─── Запуск ──────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(os.environ["TELEGRAM_BOT_TOKEN"]).build()

    # Команды
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("new", cmd_new))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("cancel", cmd_cancel))
    app.add_handler(CommandHandler("broadcast", cmd_broadcast))
    app.add_handler(CommandHandler("users", cmd_users))
    app.add_handler(CommandHandler("adduser", cmd_adduser))
    app.add_handler(CommandHandler("removeuser", cmd_removeuser))

    # Все callbacks — глобальные, работают из любого места
    app.add_handler(CallbackQueryHandler(cb_menu,             pattern="^menu_"))
    app.add_handler(CallbackQueryHandler(cb_direction,        pattern="^dir_"))
    app.add_handler(CallbackQueryHandler(cb_hasdoc,           pattern="^hasdoc_"))
    app.add_handler(CallbackQueryHandler(cb_answer,           pattern="^ans_"))
    app.add_handler(CallbackQueryHandler(cb_neg,              pattern="^neg_"))
    app.add_handler(CallbackQueryHandler(cb_criteria_q,       pattern="^(yes|no)_criteria$"))
    app.add_handler(CallbackQueryHandler(cb_review,           pattern="^review_"))
    app.add_handler(CallbackQueryHandler(cb_fmt,              pattern="^fmt_"))
    app.add_handler(CallbackQueryHandler(cb_letter_type,      pattern="^letter_"))
    app.add_handler(CallbackQueryHandler(cb_doc_action,       pattern="^docact_"))
    app.add_handler(CallbackQueryHandler(cb_photo_action,     pattern="^photo_act_"))
    app.add_handler(CallbackQueryHandler(cb_save_to_word,     pattern="^save_to_word$"))
    app.add_handler(CallbackQueryHandler(cb_free_tz_nodocs,   pattern="^free_tz_no_docs$"))

    # Сообщения — один роутер
    app.add_handler(MessageHandler(filters.PHOTO, chat_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, chat_handler))
    app.add_handler(MessageHandler((filters.TEXT | filters.VOICE) & ~filters.COMMAND, chat_handler))

    logger.info("Bot v13 started")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
